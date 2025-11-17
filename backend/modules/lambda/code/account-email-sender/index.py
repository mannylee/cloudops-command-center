import boto3
import json
import os
from datetime import datetime, timezone
from io import BytesIO
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import traceback

# Environment variables
S3_BUCKET_NAME = os.environ.get('S3_BUCKET_NAME')
S3_ATTACHMENTS_PREFIX = os.environ.get('S3_ATTACHMENTS_PREFIX', 'email-attachments/account-emails')
SENDER_EMAIL = os.environ.get('SENDER_EMAIL')
ACCOUNT_EMAIL_CC = os.environ.get('ACCOUNT_EMAIL_CC', '')
PRESIGNED_URL_EXPIRATION = int(os.environ.get('PRESIGNED_URL_EXPIRATION', '604800'))  # 7 days
EMAIL_ATTACHMENT_SIZE_THRESHOLD_MB = float(os.environ.get('EMAIL_ATTACHMENT_SIZE_THRESHOLD_MB', '5'))
LOG_LEVEL = os.environ.get('LOG_LEVEL', 'INFO')

# Initialize AWS clients
s3_client = boto3.client('s3')
ses_client = boto3.client('ses')

def lambda_handler(event, context):
    """
    Process SQS messages in batch for account-specific email sending
    Returns: batchItemFailures for partial batch failure handling
    """
    print(f"Received event with {len(event.get('Records', []))} messages")
    
    batch_item_failures = []
    
    for record in event.get('Records', []):
        try:
            message_body = json.loads(record['body'])
            print(f"Processing message for email: {message_body.get('ownerEmail')}")
            
            # Process the account email
            process_account_email(message_body)
            
        except Exception as e:
            print(f"Error processing message {record['messageId']}: {str(e)}")
            traceback.print_exc()
            
            # Add to batch item failures for retry
            batch_item_failures.append({
                'itemIdentifier': record['messageId']
            })
    
    print(f"Batch processing complete: {len(batch_item_failures)} failures")
    
    return {
        'batchItemFailures': batch_item_failures
    }


def process_account_email(message_body):
    """
    Process a single account email message
    Parse SQS message, generate Excel, upload to S3, send email
    """
    try:
        # Parse message
        account_ids = message_body.get('accountIds', [])
        account_names = message_body.get('accountNames', [])
        owner_email = message_body.get('ownerEmail')
        is_consolidated = message_body.get('isConsolidated', False)
        events = message_body.get('events', [])
        email_mappings_info = message_body.get('emailMappingsInfo', [])
        
        print(f"Processing account email for: {owner_email}")
        print(f"Accounts: {', '.join(account_ids)}")
        print(f"Events: {len(events)}")
        print(f"Consolidated: {is_consolidated}")
        
        # Generate Excel report
        excel_bytes = create_account_excel_report(events, account_ids, account_names, email_mappings_info)
        
        # Upload to S3 and get presigned URL
        presigned_url = upload_to_s3_with_partitioning(excel_bytes, owner_email)
        
        # Check file size
        file_size_mb = check_file_size(excel_bytes)
        
        # Determine attachment strategy
        attach_file = should_attach_file(file_size_mb)
        
        # Generate HTML email
        html_content = generate_account_summary_html(
            events, 
            account_ids, 
            account_names, 
            presigned_url, 
            is_consolidated, 
            file_size_mb, 
            attach_file
        )
        
        # Get CC email if configured
        cc_email = ACCOUNT_EMAIL_CC if ACCOUNT_EMAIL_CC else None
        
        # Send email (with or without attachment)
        username = extract_username_from_email(owner_email)
        filename = f"AWS_Health_Events_{username}_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        if attach_file:
            send_account_email_with_attachment(
                html_content, 
                owner_email, 
                account_ids, 
                account_names, 
                excel_bytes, 
                filename, 
                cc_email
            )
        else:
            send_account_email_link_only(
                html_content, 
                owner_email, 
                account_ids, 
                account_names, 
                cc_email
            )
        
        print(f"Account email processed successfully for: {owner_email}")
        
    except Exception as e:
        print(f"Error processing account email: {str(e)}")
        traceback.print_exc()
        raise


def generate_summary_sheet_data(events, account_ids=None):
    """
    Calculate summary statistics for Summary sheet
    Returns: dict with total_events, events_by_risk, events_by_service, events_by_category, events_by_region
    """
    try:
        print(f"Generating summary statistics for {len(events)} events")
        
        # Calculate total events
        total_events = len(events)
        
        # Group by risk level
        events_by_risk = defaultdict(int)
        for event in events:
            risk_level = event.get('riskLevel', 'Unknown')
            events_by_risk[risk_level] += 1
        
        # Group by service
        events_by_service = defaultdict(int)
        for event in events:
            service = event.get('service', 'Unknown')
            events_by_service[service] += 1
        
        # Group by category
        events_by_category = defaultdict(int)
        for event in events:
            category = event.get('eventTypeCategory', 'Unknown')
            events_by_category[category] += 1
        
        # Group by region
        events_by_region = defaultdict(int)
        for event in events:
            region = event.get('region', 'Unknown')
            events_by_region[region] += 1
        
        # Count unique accounts if not provided
        if account_ids is None:
            unique_accounts = set()
            for event in events:
                account_id = event.get('accountId')
                if account_id:
                    unique_accounts.add(account_id)
            total_accounts = len(unique_accounts)
        else:
            total_accounts = len(account_ids)
        
        summary_data = {
            'total_events': total_events,
            'total_accounts': total_accounts,
            'events_by_risk': dict(events_by_risk),
            'events_by_service': dict(events_by_service),
            'events_by_category': dict(events_by_category),
            'events_by_region': dict(events_by_region)
        }
        
        print(f"Summary: {total_events} events, {total_accounts} accounts")
        return summary_data
        
    except Exception as e:
        print(f"Error generating summary sheet data: {str(e)}")
        traceback.print_exc()
        raise



def create_summary_sheet(workbook, summary_data, account_ids, account_names, is_master=False):
    """
    Create Summary sheet as first sheet in Excel workbook
    Include overall statistics and breakdowns by risk level, service, category, region
    """
    try:
        print("Creating Summary sheet...")
        
        # Create or get the first sheet
        if 'Sheet' in workbook.sheetnames:
            ws = workbook['Sheet']
            ws.title = 'Summary'
        else:
            ws = workbook.create_sheet('Summary', 0)
        
        # Header styles
        header_font = Font(bold=True, size=14)
        section_font = Font(bold=True, size=12)
        table_header_font = Font(bold=True, color="FFFFFF")
        table_header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        current_row = 1
        
        # Title
        ws.cell(row=current_row, column=1).value = "AWS Health Events Summary"
        ws.cell(row=current_row, column=1).font = header_font
        current_row += 1
        
        # Generated timestamp
        ws.cell(row=current_row, column=1).value = f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')}"
        current_row += 2
        
        # Account information
        if not is_master:
            if len(account_ids) == 1:
                ws.cell(row=current_row, column=1).value = f"Account: {account_names[0]}"
                current_row += 1
                ws.cell(row=current_row, column=1).value = f"Account ID: {account_ids[0]}"
            else:
                ws.cell(row=current_row, column=1).value = "Accounts:"
                current_row += 1
                for acc_id, acc_name in zip(account_ids, account_names):
                    ws.cell(row=current_row, column=1).value = f"  {acc_name} ({acc_id})"
                    current_row += 1
            current_row += 1
        
        # Overall Statistics
        ws.cell(row=current_row, column=1).value = "OVERALL STATISTICS"
        ws.cell(row=current_row, column=1).font = section_font
        current_row += 1
        
        ws.cell(row=current_row, column=1).value = "Total Events:"
        ws.cell(row=current_row, column=2).value = summary_data['total_events']
        current_row += 1
        
        if is_master:
            ws.cell(row=current_row, column=1).value = "Total Accounts:"
            ws.cell(row=current_row, column=2).value = summary_data['total_accounts']
            current_row += 1
        
        current_row += 1
        
        # Events by Risk Level
        ws.cell(row=current_row, column=1).value = "EVENTS BY RISK LEVEL"
        ws.cell(row=current_row, column=1).font = section_font
        current_row += 1
        
        # Table header
        ws.cell(row=current_row, column=1).value = "Risk Level"
        ws.cell(row=current_row, column=2).value = "Count"
        ws.cell(row=current_row, column=1).font = table_header_font
        ws.cell(row=current_row, column=2).font = table_header_font
        ws.cell(row=current_row, column=1).fill = table_header_fill
        ws.cell(row=current_row, column=2).fill = table_header_fill
        current_row += 1
        
        # Sort by risk level priority
        risk_order = {'CRITICAL': 0, 'HIGH': 1, 'MEDIUM': 2, 'LOW': 3}
        for risk_level, count in sorted(summary_data['events_by_risk'].items(), 
                                       key=lambda x: risk_order.get(x[0], 99)):
            ws.cell(row=current_row, column=1).value = risk_level
            ws.cell(row=current_row, column=2).value = count
            current_row += 1
        
        current_row += 1
        
        # Events by Service
        ws.cell(row=current_row, column=1).value = "EVENTS BY SERVICE"
        ws.cell(row=current_row, column=1).font = section_font
        current_row += 1
        
        ws.cell(row=current_row, column=1).value = "Service"
        ws.cell(row=current_row, column=2).value = "Count"
        ws.cell(row=current_row, column=1).font = table_header_font
        ws.cell(row=current_row, column=2).font = table_header_font
        ws.cell(row=current_row, column=1).fill = table_header_fill
        ws.cell(row=current_row, column=2).fill = table_header_fill
        current_row += 1
        
        for service, count in sorted(summary_data['events_by_service'].items(), 
                                    key=lambda x: x[1], reverse=True)[:15]:
            ws.cell(row=current_row, column=1).value = service
            ws.cell(row=current_row, column=2).value = count
            current_row += 1
        
        current_row += 1
        
        # Events by Category
        ws.cell(row=current_row, column=1).value = "EVENTS BY CATEGORY"
        ws.cell(row=current_row, column=1).font = section_font
        current_row += 1
        
        ws.cell(row=current_row, column=1).value = "Category"
        ws.cell(row=current_row, column=2).value = "Count"
        ws.cell(row=current_row, column=1).font = table_header_font
        ws.cell(row=current_row, column=2).font = table_header_font
        ws.cell(row=current_row, column=1).fill = table_header_fill
        ws.cell(row=current_row, column=2).fill = table_header_fill
        current_row += 1
        
        for category, count in sorted(summary_data['events_by_category'].items(), 
                                     key=lambda x: x[1], reverse=True):
            ws.cell(row=current_row, column=1).value = category
            ws.cell(row=current_row, column=2).value = count
            current_row += 1
        
        current_row += 1
        
        # Events by Region
        ws.cell(row=current_row, column=1).value = "EVENTS BY REGION"
        ws.cell(row=current_row, column=1).font = section_font
        current_row += 1
        
        ws.cell(row=current_row, column=1).value = "Region"
        ws.cell(row=current_row, column=2).value = "Count"
        ws.cell(row=current_row, column=1).font = table_header_font
        ws.cell(row=current_row, column=2).font = table_header_font
        ws.cell(row=current_row, column=1).fill = table_header_fill
        ws.cell(row=current_row, column=2).fill = table_header_fill
        current_row += 1
        
        for region, count in sorted(summary_data['events_by_region'].items(), 
                                   key=lambda x: x[1], reverse=True)[:15]:
            ws.cell(row=current_row, column=1).value = region
            ws.cell(row=current_row, column=2).value = count
            current_row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        
        print("Summary sheet created successfully")
        
    except Exception as e:
        print(f"Error creating summary sheet: {str(e)}")
        traceback.print_exc()
        raise



def create_health_events_sheet(workbook, events, account_ids):
    """
    Create Health Events sheet with 15 columns
    Include only events for specified account(s)
    Sort by account, then by event date
    """
    try:
        print(f"Creating Health Events sheet with {len(events)} events...")
        
        ws = workbook.create_sheet('Health Events')
        
        # Define headers (15 columns)
        headers = [
            'Event ARN', 'Service', 'Event Type', 'Category', 'Region',
            'Status', 'Start Time', 'Last Updated', 'Account ID',
            'Risk Level', 'Risk Category', 'Time Sensitivity',
            'Affected Resources', 'Description', 'Required Actions'
        ]
        
        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Set column widths
        column_widths = [50, 20, 30, 20, 15, 15, 20, 20, 15, 15, 20, 20, 30, 60, 60]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width
        
        # Filter events for specified accounts
        filtered_events = [e for e in events if e.get('accountId') in account_ids]
        
        # Sort by account ID, then by lastUpdateTime
        sorted_events = sorted(filtered_events, 
                              key=lambda x: (x.get('accountId', ''), x.get('lastUpdateTime', '')))
        
        # Write event data
        for row_num, event in enumerate(sorted_events, 2):
            ws.cell(row=row_num, column=1).value = event.get('eventArn', 'N/A')
            ws.cell(row=row_num, column=2).value = event.get('service', 'N/A')
            ws.cell(row=row_num, column=3).value = event.get('eventType', 'N/A')
            ws.cell(row=row_num, column=4).value = event.get('eventTypeCategory', 'N/A')
            ws.cell(row=row_num, column=5).value = event.get('region', 'N/A')
            ws.cell(row=row_num, column=6).value = event.get('statusCode', 'N/A')
            ws.cell(row=row_num, column=7).value = event.get('startTime', 'N/A')
            ws.cell(row=row_num, column=8).value = event.get('lastUpdateTime', 'N/A')
            ws.cell(row=row_num, column=9).value = event.get('accountId', 'N/A')
            ws.cell(row=row_num, column=10).value = event.get('riskLevel', 'N/A')
            ws.cell(row=row_num, column=11).value = event.get('riskCategory', 'N/A')
            ws.cell(row=row_num, column=12).value = event.get('timeSensitivity', 'N/A')
            ws.cell(row=row_num, column=13).value = event.get('affectedResources', 'N/A')
            ws.cell(row=row_num, column=14).value = event.get('description', 'N/A')
            ws.cell(row=row_num, column=15).value = event.get('requiredActions', 'N/A')
            
            # Set default row height (15 is Excel's default)
            ws.row_dimensions[row_num].height = 15
        
        print(f"Health Events sheet created with {len(sorted_events)} events")
        
    except Exception as e:
        print(f"Error creating health events sheet: {str(e)}")
        traceback.print_exc()
        raise



def create_account_email_mapping_sheet(workbook, email_mappings_info, is_master=False):
    """
    Create Account Email Mapping sheet
    For account emails: columns are Account ID, Email Address, Mapping Source (no Availability Status)
    For master email: columns include Availability Status
    List only accounts included in this email (for account emails)
    """
    try:
        print("Creating Account Email Mapping sheet...")
        
        ws = workbook.create_sheet('Account Email Mapping')
        
        # Define headers based on email type
        if is_master:
            headers = ['Account ID', 'Account Name', 'Email Address', 'Mapping Source', 'Availability Status']
        else:
            headers = ['Account ID', 'Account Name', 'Email Address', 'Mapping Source']
        
        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Set column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 25
        if is_master:
            ws.column_dimensions['E'].width = 25
        
        # Write mapping data
        for row_num, mapping in enumerate(email_mappings_info, 2):
            ws.cell(row=row_num, column=1).value = mapping.get('accountId', 'N/A')
            ws.cell(row=row_num, column=2).value = mapping.get('accountName', 'N/A')
            ws.cell(row=row_num, column=3).value = mapping.get('emailAddress', 'N/A')
            ws.cell(row=row_num, column=4).value = mapping.get('mappingSource', 'N/A')
            if is_master:
                ws.cell(row=row_num, column=5).value = mapping.get('availabilityStatus', 'N/A')
        
        print(f"Account Email Mapping sheet created with {len(email_mappings_info)} accounts")
        
    except Exception as e:
        print(f"Error creating account email mapping sheet: {str(e)}")
        traceback.print_exc()
        raise



def create_account_excel_report(events, account_ids, account_names, email_mappings_info):
    """
    Generate complete Excel workbook with all 3 sheets
    Sheet order: Summary, Health Events, Account Email Mapping
    Returns: Excel file as bytes
    """
    try:
        print(f"Creating Excel report for accounts: {', '.join(account_ids)}")
        
        # Create workbook
        wb = Workbook()
        
        # Generate summary data
        summary_data = generate_summary_sheet_data(events, account_ids)
        
        # Create Summary sheet (first sheet)
        create_summary_sheet(wb, summary_data, account_ids, account_names, is_master=False)
        
        # Create Health Events sheet
        create_health_events_sheet(wb, events, account_ids)
        
        # Create Account Email Mapping sheet
        create_account_email_mapping_sheet(wb, email_mappings_info, is_master=False)
        
        # Remove default sheet if it exists
        if 'Sheet' in wb.sheetnames and len(wb.sheetnames) > 1:
            wb.remove(wb['Sheet'])
        
        # Save to BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        print("Excel report created successfully")
        return excel_buffer.getvalue()
        
    except Exception as e:
        print(f"Error creating Excel report: {str(e)}")
        traceback.print_exc()
        raise



def extract_username_from_email(email):
    """
    Extract username portion from email address (before @ symbol)
    Handle plus addressing (user+tag@domain.com)
    Handle dots in username (john.doe@domain.com)
    Returns: username string
    """
    try:
        if not email or '@' not in email:
            return 'unknown'
        
        # Extract everything before @
        username = email.split('@')[0]
        
        # Username can contain dots and plus signs, which are valid
        # We keep them as-is for the S3 path
        
        print(f"Extracted username '{username}' from email '{email}'")
        return username
        
    except Exception as e:
        print(f"Error extracting username from email '{email}': {str(e)}")
        return 'unknown'



def upload_to_s3_with_partitioning(excel_bytes, recipient_email):
    """
    Upload Excel file to S3 with username-at-day-level partitioning
    Generate S3 key: account-emails/YYYY/MM/DD/{username}/AWS_Health_Events_{username}_{timestamp}.xlsx
    Generate presigned URL with 7-day expiration
    Returns: presigned URL
    """
    try:
        # Extract username from email
        username = extract_username_from_email(recipient_email)
        
        # Generate date partitioning
        now = datetime.now(timezone.utc)
        year = now.strftime('%Y')
        month = now.strftime('%m')
        day = now.strftime('%d')
        timestamp = now.strftime('%Y%m%d_%H%M%S')
        
        # Generate filename with username
        filename = f"AWS_Health_Events_{username}_{timestamp}.xlsx"
        
        # Generate S3 key with username-at-day-level partitioning
        # Format: account-emails/YYYY/MM/DD/{username}/filename.xlsx
        key = f"{S3_ATTACHMENTS_PREFIX}/{year}/{month}/{day}/{username}/{filename}"
        
        print(f"Uploading to S3: s3://{S3_BUCKET_NAME}/{key}")
        
        # Upload to S3
        s3_client.put_object(
            Bucket=S3_BUCKET_NAME,
            Key=key,
            Body=excel_bytes,
            ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            ServerSideEncryption='AES256'
        )
        
        # Generate presigned URL
        presigned_url = s3_client.generate_presigned_url(
            'get_object',
            Params={
                'Bucket': S3_BUCKET_NAME,
                'Key': key
            },
            ExpiresIn=PRESIGNED_URL_EXPIRATION
        )
        
        print(f"File uploaded successfully, presigned URL generated (expires in {PRESIGNED_URL_EXPIRATION} seconds)")
        return presigned_url
        
    except Exception as e:
        print(f"Error uploading to S3: {str(e)}")
        traceback.print_exc()
        raise



def check_file_size(excel_file_bytes):
    """
    Calculate Excel file size in MB
    Returns: file size as float (MB)
    """
    try:
        size_bytes = len(excel_file_bytes)
        size_mb = size_bytes / (1024 * 1024)
        
        print(f"File size: {size_mb:.2f} MB ({size_bytes} bytes)")
        return size_mb
        
    except Exception as e:
        print(f"Error checking file size: {str(e)}")
        return 0.0



def should_attach_file(file_size_mb):
    """
    Determine whether to attach file or link-only based on size threshold
    Check if file size < 5 MB threshold
    Log decision for monitoring
    Returns: True if file should be attached, False for link-only
    """
    try:
        attach = file_size_mb < EMAIL_ATTACHMENT_SIZE_THRESHOLD_MB
        
        if attach:
            print(f"File size {file_size_mb:.2f} MB < {EMAIL_ATTACHMENT_SIZE_THRESHOLD_MB} MB threshold: ATTACHING file")
        else:
            print(f"File size {file_size_mb:.2f} MB >= {EMAIL_ATTACHMENT_SIZE_THRESHOLD_MB} MB threshold: LINK ONLY")
        
        return attach
        
    except Exception as e:
        print(f"Error determining attachment decision: {str(e)}")
        # Default to link-only on error
        return False



def generate_account_summary_html(events, account_ids, account_names, presigned_url, is_consolidated, file_size_mb, attachment_included):
    """
    Generate HTML email body with statistics
    Include comma-separated account IDs
    Adjust messaging based on attachment vs link-only
    """
    try:
        print("Generating HTML email content...")
        
        # Generate summary data
        summary_data = generate_summary_sheet_data(events, account_ids)
        
        # Current date
        current_date = datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')
        
        # Account information
        if is_consolidated:
            account_info = f"<p><strong>Account ID(s):</strong> {', '.join(account_ids)}</p>"
            title = "AWS Health Events for Multiple Accounts"
        else:
            account_info = f"<p><strong>Account ID(s):</strong> {account_ids[0]}</p>"
            title = f"AWS Health Events for Account: {account_names[0]}"
        
        # Attachment messaging
        if attachment_included:
            attachment_msg = """
                <p>The Excel report is attached to this email.</p>
                <p>You can also download it here: <a href="{url}">AWS Health Events Report</a></p>
                <p><small>This link will expire in 7 days.</small></p>
            """.format(url=presigned_url)
        else:
            attachment_msg = """
                <p><strong>Due to the report size, the Excel file is only available via the download link below.</strong></p>
                <p>Download the full report: <a href="{url}">AWS Health Events Report</a></p>
                <p><small>This link will expire in 7 days.</small></p>
            """.format(url=presigned_url)
        
        # Build HTML
        html = f"""
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; }}
                .header {{ background-color: #232F3E; color: white; padding: 20px; }}
                .content {{ padding: 20px; }}
                .summary {{ margin: 20px 0; }}
                .stats {{ background-color: #f5f5f5; padding: 15px; border-radius: 5px; margin: 10px 0; }}
                .download-button {{
                    display: inline-block;
                    background-color: #FF9900;
                    color: white;
                    padding: 12px 24px;
                    text-decoration: none;
                    border-radius: 4px;
                    font-weight: bold;
                    margin: 20px 0;
                }}
                .download-button:hover {{ background-color: #EC7211; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>{title}</h1>
                <p>Generated: {current_date}</p>
            </div>
            <div class="content">
                <div class="summary">
                    <h2>Summary</h2>
                    {account_info}
                    <p><strong>Total Events:</strong> {summary_data['total_events']}</p>
                    
                    <div class="stats">
                        <h3>Events by Risk Level</h3>
                        <ul>
        """
        
        # Sort risk levels by severity
        risk_order = {'CRITICAL': 0, 'HIGH': 1, 'MEDIUM': 2, 'LOW': 3, 'Unknown': 4}
        for risk_level, count in sorted(summary_data['events_by_risk'].items(), 
                                       key=lambda x: risk_order.get(x[0], 99)):
            if risk_level == 'CRITICAL':
                html += f'<li><strong style="color: #ff4444;">🔴 CRITICAL: {count}</strong></li>\n'
            elif risk_level == 'HIGH':
                html += f'<li><strong style="color: #ff9900;">🟠 HIGH: {count}</strong></li>\n'
            elif risk_level == 'MEDIUM':
                html += f'<li><strong style="color: #ffcc00;">🟡 MEDIUM: {count}</strong></li>\n'
            elif risk_level == 'LOW':
                html += f'<li>🟢 LOW: {count}</li>\n'
            else:
                html += f'<li>⚪ {risk_level}: {count}</li>\n'
        
        html += """
                        </ul>
                    </div>
                    
                    <div class="stats">
                        <h3>Events by Service</h3>
                        <ul>
        """
        
        for service, count in sorted(summary_data['events_by_service'].items(), 
                                    key=lambda x: x[1], reverse=True)[:10]:
            html += f"<li>{service}: {count}</li>\n"
        
        html += """
                        </ul>
                    </div>
                    
                    <div class="stats">
                        <h3>Events by Category</h3>
                        <ul>
        """
        
        for category, count in sorted(summary_data['events_by_category'].items(), 
                                     key=lambda x: x[1], reverse=True):
            html += f"<li>{category}: {count}</li>\n"
        
        html += """
                        </ul>
                    </div>
                    
                    <div class="stats">
                        <h3>Events by Region</h3>
                        <ul>
        """
        
        for region, count in sorted(summary_data['events_by_region'].items(), 
                                   key=lambda x: x[1], reverse=True)[:10]:
            html += f"<li>{region}: {count}</li>\n"
        
        html += f"""
                        </ul>
                    </div>
                </div>
                
                <div class="summary">
                    <h2>Download Report</h2>
                    {attachment_msg}
                </div>
            </div>
        </body>
        </html>
        """
        
        print("HTML email content generated successfully")
        return html
        
    except Exception as e:
        print(f"Error generating HTML summary: {str(e)}")
        traceback.print_exc()
        raise



def send_account_email_with_attachment(html_content, recipient_email, account_ids, account_names, excel_file_bytes, filename, cc_email=None):
    """
    Send email via SES with Excel file attached using SendRawEmail API
    Include subject line with recipient username
    Optionally include CC recipient if configured
    """
    try:
        # Extract username for subject line
        username = extract_username_from_email(recipient_email)
        
        # Create subject line with username
        current_date = datetime.now(timezone.utc).strftime('%Y-%m-%d')
        subject = f"AWS Health Events Summary [{username}] - {current_date}"
        
        print(f"Sending email with attachment to: {recipient_email}")
        if cc_email:
            print(f"CC: {cc_email}")
        
        # Create MIME message
        msg = MIMEMultipart('mixed')
        msg['Subject'] = subject
        msg['From'] = SENDER_EMAIL
        msg['To'] = recipient_email
        if cc_email:
            msg['Cc'] = cc_email
        
        # Add HTML body
        body = MIMEMultipart('alternative')
        html_part = MIMEText(html_content, 'html')
        body.attach(html_part)
        msg.attach(body)
        
        # Add Excel attachment
        attachment = MIMEApplication(excel_file_bytes)
        attachment.add_header('Content-Disposition', 'attachment', filename=filename)
        msg.attach(attachment)
        
        # Send via SES
        destinations = [recipient_email]
        if cc_email:
            destinations.append(cc_email)
        
        response = ses_client.send_raw_email(
            Source=SENDER_EMAIL,
            Destinations=destinations,
            RawMessage={
                'Data': msg.as_string()
            }
        )
        
        print(f"Email with attachment sent successfully. MessageId: {response['MessageId']}")
        
    except Exception as e:
        print(f"Error sending email with attachment: {str(e)}")
        traceback.print_exc()
        raise



def send_account_email_link_only(html_content, recipient_email, account_ids, account_names, cc_email=None):
    """
    Send email via SES without attachment (link only) using SendEmail API
    Include subject line with username
    Optionally include CC recipient if configured
    """
    try:
        # Extract username for subject line
        username = extract_username_from_email(recipient_email)
        
        # Create subject line with username
        current_date = datetime.now(timezone.utc).strftime('%Y-%m-%d')
        subject = f"AWS Health Events Summary [{username}] - {current_date}"
        
        print(f"Sending email (link only) to: {recipient_email}")
        if cc_email:
            print(f"CC: {cc_email}")
        
        # Prepare destination
        destination = {'ToAddresses': [recipient_email]}
        if cc_email:
            destination['CcAddresses'] = [cc_email]
        
        # Send via SES
        response = ses_client.send_email(
            Source=SENDER_EMAIL,
            Destination=destination,
            Message={
                'Subject': {
                    'Data': subject,
                    'Charset': 'UTF-8'
                },
                'Body': {
                    'Html': {
                        'Data': html_content,
                        'Charset': 'UTF-8'
                    }
                }
            }
        )
        
        print(f"Email (link only) sent successfully. MessageId: {response['MessageId']}")
        
    except Exception as e:
        print(f"Error sending email (link only): {str(e)}")
        traceback.print_exc()
        raise
