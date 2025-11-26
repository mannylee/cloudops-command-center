import boto3
import json
import os
from datetime import datetime, timezone
from io import BytesIO
from collections import defaultdict
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import traceback

# Environment variables
DYNAMODB_HEALTH_EVENTS_TABLE_NAME = os.environ.get('DYNAMODB_HEALTH_EVENTS_TABLE_NAME')
SENDER_EMAIL = os.environ.get('SENDER_EMAIL')
MASTER_RECIPIENT_EMAIL = os.environ.get('MASTER_RECIPIENT_EMAIL')
S3_BUCKET_NAME = os.environ.get('S3_BUCKET_NAME')
S3_ATTACHMENTS_PREFIX = os.environ.get('S3_ATTACHMENTS_PREFIX', 'email-attachments')
PRESIGNED_URL_EXPIRATION = int(os.environ.get('PRESIGNED_URL_EXPIRATION', '604800'))  # 7 days
LOG_LEVEL = os.environ.get('LOG_LEVEL', 'INFO')
ACCOUNT_EMAIL_MAPPINGS_TABLE = os.environ.get('ACCOUNT_EMAIL_MAPPINGS_TABLE', '')
ACCOUNT_EMAIL_QUEUE_URL = os.environ.get('ACCOUNT_EMAIL_QUEUE_URL', '')
ENABLE_PER_ACCOUNT_EMAILS = os.environ.get('ENABLE_PER_ACCOUNT_EMAILS', 'true').lower() == 'true'

# Initialize AWS clients
dynamodb = boto3.resource('dynamodb')
s3_client = boto3.client('s3')
ses_client = boto3.client('ses')
organizations_client = boto3.client('organizations')
sqs_client = boto3.client('sqs')


def convert_decimal_to_number(obj):
    """
    Recursively convert Decimal objects to int or float for JSON serialization
    DynamoDB returns numbers as Decimal objects which aren't JSON serializable
    """
    if isinstance(obj, list):
        return [convert_decimal_to_number(item) for item in obj]
    elif isinstance(obj, dict):
        return {key: convert_decimal_to_number(value) for key, value in obj.items()}
    elif isinstance(obj, Decimal):
        # Convert to int if it's a whole number, otherwise float
        if obj % 1 == 0:
            return int(obj)
        else:
            return float(obj)
    else:
        return obj


def lambda_handler(event, context):
    """
    Lambda handler for processing email notification requests
    Triggered directly by EventBridge on a schedule
    """
    try:
        print(f"Email processor triggered by EventBridge")
        print(f"Event: {json.dumps(event)}")
        
        # Generate and send summary email
        generate_and_send_summary_email()
        
        return {
            'statusCode': 200,
            'body': json.dumps('Email processing completed successfully')
        }
        
    except Exception as e:
        print(f"Error in lambda_handler: {str(e)}")
        traceback.print_exc()
        raise


def generate_and_send_summary_email():
    """
    Generate and send summary email with open health events
    Implements hybrid attachment logic: < 5MB attach, >= 5MB link-only
    """
    try:
        print("Fetching open health events from DynamoDB...")
        
        # Query DynamoDB for open/unresolved events
        events = fetch_open_health_events()
        
        if not events:
            print("No open health events found")
            # Still send email with "no events" message
            send_no_events_email()
            return
        
        print(f"Found {len(events)} open health events")
        
        # Generate Excel report with Summary sheet
        excel_buffer = create_excel_report(events)
        excel_bytes = excel_buffer.getvalue()
        
        # Check file size
        file_size_bytes = len(excel_bytes)
        file_size_mb = file_size_bytes / (1024 * 1024)
        print(f"Excel file size: {file_size_mb:.2f} MB ({file_size_bytes} bytes)")
        
        # Determine attachment strategy (5 MB threshold)
        attach_file = file_size_mb < 5.0
        
        if attach_file:
            print(f"File size {file_size_mb:.2f} MB < 5 MB threshold: ATTACHING file")
        else:
            print(f"File size {file_size_mb:.2f} MB >= 5 MB threshold: LINK ONLY")
        
        # Upload to S3 and get presigned URL (always upload regardless of attachment decision)
        attachment_key = upload_attachment_to_s3(excel_buffer)
        presigned_url = generate_presigned_url(attachment_key)
        
        # Generate HTML email content
        html_content = generate_master_summary_html(events, presigned_url, attach_file)
        
        # Send email with or without attachment
        if attach_file:
            send_master_email_with_attachment(html_content, presigned_url, excel_bytes)
        else:
            send_master_email_link_only(html_content, presigned_url)
        
        print("Master email sent successfully")
        
        # Process per-account emails after master email is sent
        try:
            process_per_account_emails()
        except Exception as e:
            print(f"Error processing per-account emails (master email succeeded): {str(e)}")
            traceback.print_exc()
            # Don't raise - master email already succeeded
        
    except Exception as e:
        print(f"Error generating summary email: {str(e)}")
        traceback.print_exc()
        raise


def fetch_open_health_events():
    """
    Fetch open/unresolved health events from DynamoDB
    """
    try:
        table = dynamodb.Table(DYNAMODB_HEALTH_EVENTS_TABLE_NAME)
        
        # Scan for events with statusCode = 'open' or 'upcoming'
        response = table.scan(
            FilterExpression='#status IN (:open, :upcoming, :scheduled)',
            ExpressionAttributeNames={
                '#status': 'statusCode'
            },
            ExpressionAttributeValues={
                ':open': 'open',
                ':upcoming': 'upcoming',
                ':scheduled': 'scheduled'
            }
        )
        
        events = response.get('Items', [])
        
        # Handle pagination
        while 'LastEvaluatedKey' in response:
            response = table.scan(
                FilterExpression='#status IN (:open, :upcoming, :scheduled)',
                ExpressionAttributeNames={
                    '#status': 'statusCode'
                },
                ExpressionAttributeValues={
                    ':open': 'open',
                    ':upcoming': 'upcoming',
                    ':scheduled': 'scheduled'
                },
                ExclusiveStartKey=response['LastEvaluatedKey']
            )
            events.extend(response.get('Items', []))
        
        print(f"Fetched {len(events)} open health events")
        return events
        
    except Exception as e:
        print(f"Error fetching health events: {str(e)}")
        traceback.print_exc()
        raise


def generate_master_summary_data(events):
    """
    Generate summary statistics for master email
    """
    try:
        total_events = len(events)
        
        # Count unique accounts
        unique_accounts = set()
        for event in events:
            account_id = event.get('accountId')
            if account_id:
                unique_accounts.add(account_id)
        total_accounts = len(unique_accounts)
        
        # Group by risk level
        events_by_risk = defaultdict(int)
        for event in events:
            risk_level = event.get('riskLevel', 'Unknown')
            events_by_risk[risk_level] += 1
        
        # Group by service
        events_by_service = defaultdict(int)
        for event in events:
            service = event.get('service', 'Unknown')
            events_by_service[service] += 1
        
        # Group by category
        events_by_category = defaultdict(int)
        for event in events:
            category = event.get('eventTypeCategory', 'Unknown')
            events_by_category[category] += 1
        
        # Group by region
        events_by_region = defaultdict(int)
        for event in events:
            region = event.get('region', 'Unknown')
            events_by_region[region] += 1
        
        return {
            'total_events': total_events,
            'total_accounts': total_accounts,
            'events_by_risk': dict(events_by_risk),
            'events_by_service': dict(events_by_service),
            'events_by_category': dict(events_by_category),
            'events_by_region': dict(events_by_region)
        }
    except Exception as e:
        print(f"Error generating summary data: {str(e)}")
        traceback.print_exc()
        raise


def create_master_summary_sheet(wb, summary_data):
    """
    Create Summary sheet for master email as first sheet
    """
    try:
        print("Creating Summary sheet for master email...")
        
        # Create Summary sheet as first sheet
        if 'Sheet' in wb.sheetnames:
            ws = wb['Sheet']
            ws.title = 'Summary'
        else:
            ws = wb.create_sheet('Summary', 0)
        
        # Header styles
        header_font = Font(bold=True, size=14)
        section_font = Font(bold=True, size=12)
        table_header_font = Font(bold=True, color="FFFFFF")
        table_header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        current_row = 1
        
        # Title
        ws.cell(row=current_row, column=1).value = "AWS Health Events Summary [MASTER]"
        ws.cell(row=current_row, column=1).font = header_font
        current_row += 1
        
        # Generated timestamp
        ws.cell(row=current_row, column=1).value = f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')}"
        current_row += 2
        
        # Overall Statistics
        ws.cell(row=current_row, column=1).value = "OVERALL STATISTICS"
        ws.cell(row=current_row, column=1).font = section_font
        current_row += 1
        
        ws.cell(row=current_row, column=1).value = "Total Events:"
        ws.cell(row=current_row, column=2).value = summary_data['total_events']
        current_row += 1
        
        ws.cell(row=current_row, column=1).value = "Total Accounts:"
        ws.cell(row=current_row, column=2).value = summary_data['total_accounts']
        current_row += 2
        
        # Events by Risk Level
        ws.cell(row=current_row, column=1).value = "EVENTS BY RISK LEVEL"
        ws.cell(row=current_row, column=1).font = section_font
        current_row += 1
        
        ws.cell(row=current_row, column=1).value = "Risk Level"
        ws.cell(row=current_row, column=2).value = "Count"
        ws.cell(row=current_row, column=1).font = table_header_font
        ws.cell(row=current_row, column=2).font = table_header_font
        ws.cell(row=current_row, column=1).fill = table_header_fill
        ws.cell(row=current_row, column=2).fill = table_header_fill
        current_row += 1
        
        risk_order = {'CRITICAL': 0, 'HIGH': 1, 'MEDIUM': 2, 'LOW': 3}
        for risk_level, count in sorted(summary_data['events_by_risk'].items(), 
                                       key=lambda x: risk_order.get(x[0], 99)):
            ws.cell(row=current_row, column=1).value = risk_level
            ws.cell(row=current_row, column=2).value = count
            current_row += 1
        
        current_row += 1
        
        # Events by Service
        ws.cell(row=current_row, column=1).value = "EVENTS BY SERVICE"
        ws.cell(row=current_row, column=1).font = section_font
        current_row += 1
        
        ws.cell(row=current_row, column=1).value = "Service"
        ws.cell(row=current_row, column=2).value = "Count"
        ws.cell(row=current_row, column=1).font = table_header_font
        ws.cell(row=current_row, column=2).font = table_header_font
        ws.cell(row=current_row, column=1).fill = table_header_fill
        ws.cell(row=current_row, column=2).fill = table_header_fill
        current_row += 1
        
        for service, count in sorted(summary_data['events_by_service'].items(), 
                                    key=lambda x: x[1], reverse=True)[:15]:
            ws.cell(row=current_row, column=1).value = service
            ws.cell(row=current_row, column=2).value = count
            current_row += 1
        
        current_row += 1
        
        # Events by Category
        ws.cell(row=current_row, column=1).value = "EVENTS BY CATEGORY"
        ws.cell(row=current_row, column=1).font = section_font
        current_row += 1
        
        ws.cell(row=current_row, column=1).value = "Category"
        ws.cell(row=current_row, column=2).value = "Count"
        ws.cell(row=current_row, column=1).font = table_header_font
        ws.cell(row=current_row, column=2).font = table_header_font
        ws.cell(row=current_row, column=1).fill = table_header_fill
        ws.cell(row=current_row, column=2).fill = table_header_fill
        current_row += 1
        
        for category, count in sorted(summary_data['events_by_category'].items(), 
                                     key=lambda x: x[1], reverse=True):
            ws.cell(row=current_row, column=1).value = category
            ws.cell(row=current_row, column=2).value = count
            current_row += 1
        
        current_row += 1
        
        # Events by Region
        ws.cell(row=current_row, column=1).value = "EVENTS BY REGION"
        ws.cell(row=current_row, column=1).font = section_font
        current_row += 1
        
        ws.cell(row=current_row, column=1).value = "Region"
        ws.cell(row=current_row, column=2).value = "Count"
        ws.cell(row=current_row, column=1).font = table_header_font
        ws.cell(row=current_row, column=2).font = table_header_font
        ws.cell(row=current_row, column=1).fill = table_header_fill
        ws.cell(row=current_row, column=2).fill = table_header_fill
        current_row += 1
        
        for region, count in sorted(summary_data['events_by_region'].items(), 
                                   key=lambda x: x[1], reverse=True)[:15]:
            ws.cell(row=current_row, column=1).value = region
            ws.cell(row=current_row, column=2).value = count
            current_row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        
        print("Summary sheet created successfully")
        
    except Exception as e:
        print(f"Error creating summary sheet: {str(e)}")
        traceback.print_exc()
        raise


def create_master_health_events_sheet(wb, events):
    """
    Create Health Events sheet for master email
    """
    try:
        print("Creating Health Events sheet...")
        
        ws = wb.create_sheet('Health Events')
        
        # Define headers (16 columns - added Account Name after Account ID)
        headers = [
            'Event ARN', 'Service', 'Event Type', 'Category', 'Region',
            'Status', 'Start Time', 'Last Updated', 'Account ID', 'Account Name',
            'Risk Level', 'Risk Category', 'Time Sensitivity',
            'Affected Resources', 'Description', 'Required Actions'
        ]
        
        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Set column widths (16 columns - added 30 for Account Name)
        column_widths = [50, 20, 30, 20, 15, 15, 20, 20, 15, 30, 15, 20, 20, 30, 60, 60]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width
        
        # Write event data
        for row_num, event in enumerate(events, 2):
            ws.cell(row=row_num, column=1).value = event.get('eventArn', 'N/A')
            ws.cell(row=row_num, column=2).value = event.get('service', 'N/A')
            ws.cell(row=row_num, column=3).value = event.get('eventType', 'N/A')
            ws.cell(row=row_num, column=4).value = event.get('eventTypeCategory', 'N/A')
            ws.cell(row=row_num, column=5).value = event.get('region', 'N/A')
            ws.cell(row=row_num, column=6).value = event.get('statusCode', 'N/A')
            ws.cell(row=row_num, column=7).value = event.get('startTime', 'N/A')
            ws.cell(row=row_num, column=8).value = event.get('lastUpdateTime', 'N/A')
            ws.cell(row=row_num, column=9).value = event.get('accountId', 'N/A')
            ws.cell(row=row_num, column=10).value = event.get('accountName', 'N/A')
            ws.cell(row=row_num, column=11).value = event.get('riskLevel', 'N/A')
            ws.cell(row=row_num, column=12).value = event.get('riskCategory', 'N/A')
            ws.cell(row=row_num, column=13).value = event.get('timeSensitivity', 'N/A')
            ws.cell(row=row_num, column=14).value = event.get('affectedResources', 'N/A')
            ws.cell(row=row_num, column=15).value = event.get('description', 'N/A')
            ws.cell(row=row_num, column=16).value = event.get('requiredActions', 'N/A')
            
            # Set default row height (15 is Excel's default)
            ws.row_dimensions[row_num].height = 15
        
        print("Health Events sheet created successfully")
        
    except Exception as e:
        print(f"Error creating health events sheet: {str(e)}")
        traceback.print_exc()
        raise


def create_master_account_mapping_sheet(wb):
    """
    Create Account Email Mapping sheet for master email with all accounts
    """
    try:
        print("Creating Account Email Mapping sheet for master email...")
        
        # Fetch email mappings info
        custom_mappings = fetch_custom_email_mappings()
        org_mappings = fetch_account_owners()
        
        if not org_mappings:
            print("No Organizations data available, skipping Account Email Mapping sheet")
            return
        
        all_accounts = merge_email_mappings(custom_mappings, org_mappings)
        email_mappings_info = build_email_mappings_info(all_accounts, custom_mappings, org_mappings)
        
        ws = wb.create_sheet('Account Email Mapping')
        
        # Define headers (master email includes Availability Status)
        headers = ['Account ID', 'Account Name', 'Email Address', 'Mapping Source', 'Availability Status']
        
        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Set column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 25
        
        # Write mapping data
        for row_num, mapping in enumerate(email_mappings_info, 2):
            ws.cell(row=row_num, column=1).value = mapping.get('accountId', 'N/A')
            ws.cell(row=row_num, column=2).value = mapping.get('accountName', 'N/A')
            ws.cell(row=row_num, column=3).value = mapping.get('emailAddress', 'N/A')
            ws.cell(row=row_num, column=4).value = mapping.get('mappingSource', 'N/A')
            ws.cell(row=row_num, column=5).value = mapping.get('availabilityStatus', 'N/A')
        
        print(f"Account Email Mapping sheet created with {len(email_mappings_info)} accounts")
        
    except Exception as e:
        print(f"Error creating account email mapping sheet (continuing without it): {str(e)}")
        traceback.print_exc()
        # Don't raise - this is optional


def create_excel_report(events):
    """
    Create Excel report with Summary, Health Events, and Account Email Mapping sheets
    """
    try:
        print("Creating master Excel report...")
        
        # Create workbook
        wb = Workbook()
        
        # Generate summary data
        summary_data = generate_master_summary_data(events)
        
        # Create Summary sheet (first sheet)
        create_master_summary_sheet(wb, summary_data)
        
        # Create Health Events sheet
        create_master_health_events_sheet(wb, events)
        
        # Create Account Email Mapping sheet
        create_master_account_mapping_sheet(wb)
        
        # Remove default sheet if it exists
        if 'Sheet' in wb.sheetnames and len(wb.sheetnames) > 1:
            wb.remove(wb['Sheet'])
        
        # Save to BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        print("Master Excel report created successfully")
        return excel_buffer
        
    except Exception as e:
        print(f"Error creating Excel report: {str(e)}")
        traceback.print_exc()
        raise


def upload_attachment_to_s3(excel_buffer):
    """
    Upload Excel attachment to S3 and return the key
    Uses master-emails prefix for master email attachments
    """
    try:
        # Generate unique filename with date partitioning
        now = datetime.now(timezone.utc)
        year = now.strftime('%Y')
        month = now.strftime('%m')
        day = now.strftime('%d')
        timestamp = now.strftime('%Y%m%d_%H%M%S')
        
        filename = f"AWS_Health_Events_Master_{timestamp}.xlsx"
        # Partition by year/month/day with master-emails prefix
        key = f"{S3_ATTACHMENTS_PREFIX}/master-emails/{year}/{month}/{day}/{filename}"
        
        print(f"Uploading master attachment to S3: s3://{S3_BUCKET_NAME}/{key}")
        
        # Upload to S3
        s3_client.put_object(
            Bucket=S3_BUCKET_NAME,
            Key=key,
            Body=excel_buffer.getvalue(),
            ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            ServerSideEncryption='AES256'
        )
        
        print("Master attachment uploaded successfully")
        return key
        
    except Exception as e:
        print(f"Error uploading attachment to S3: {str(e)}")
        traceback.print_exc()
        raise


def generate_presigned_url(key):
    """
    Generate presigned URL for S3 object
    """
    try:
        url = s3_client.generate_presigned_url(
            'get_object',
            Params={
                'Bucket': S3_BUCKET_NAME,
                'Key': key
            },
            ExpiresIn=PRESIGNED_URL_EXPIRATION
        )
        
        print(f"Generated presigned URL (expires in {PRESIGNED_URL_EXPIRATION} seconds)")
        return url
        
    except Exception as e:
        print(f"Error generating presigned URL: {str(e)}")
        traceback.print_exc()
        raise


def generate_summary_html(events, presigned_url):
    """
    Generate HTML summary for email
    """
    try:
        # Calculate statistics
        total_events = len(events)
        
        # Group by service
        services = defaultdict(int)
        for event in events:
            service = event.get('service', 'Unknown')
            services[service] += 1
        
        # Group by category
        categories = defaultdict(int)
        for event in events:
            category = event.get('eventTypeCategory', 'Unknown')
            categories[category] += 1
        
        # Group by region
        regions = defaultdict(int)
        for event in events:
            region = event.get('region', 'Unknown')
            regions[region] += 1
        
        # Group by risk level
        risk_levels = defaultdict(int)
        for event in events:
            risk_level = event.get('riskLevel', 'Unknown')
            risk_levels[risk_level] += 1
        
        # Current date
        current_date = datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')
        
        # Build HTML
        html = f"""
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; }}
                .header {{ background-color: #232F3E; color: white; padding: 20px; }}
                .content {{ padding: 20px; }}
                .summary {{ margin: 20px 0; }}
                .stats {{ background-color: #f5f5f5; padding: 15px; border-radius: 5px; margin: 10px 0; }}
                .download-button {{
                    display: inline-block;
                    background-color: #FF9900;
                    color: white;
                    padding: 12px 24px;
                    text-decoration: none;
                    border-radius: 4px;
                    font-weight: bold;
                    margin: 20px 0;
                }}
                .download-button:hover {{ background-color: #EC7211; }}
                table {{ border-collapse: collapse; width: 100%; margin-top: 20px; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #f2f2f2; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>AWS Health Events Summary Report</h1>
                <p>Generated: {current_date}</p>
            </div>
            <div class="content">
                <div class="summary">
                    <h2>Summary</h2>
                    <p><strong>Total Open/Upcoming Events:</strong> {total_events}</p>
                    
                    <div class="stats">
                        <h3>Events by Risk Level</h3>
                        <ul>
        """
        
        # Sort risk levels by severity and display with color indicators
        risk_order = {'CRITICAL': 0, 'HIGH': 1, 'MEDIUM': 2, 'LOW': 3, 'Unknown': 4}
        for risk_level, count in sorted(risk_levels.items(), key=lambda x: risk_order.get(x[0], 99)):
            # Add visual indicators for risk levels
            if risk_level == 'CRITICAL':
                html += f'<li><strong style="color: #ff4444;">🔴 CRITICAL: {count}</strong></li>\n'
            elif risk_level == 'HIGH':
                html += f'<li><strong style="color: #ff9900;">🟠 HIGH: {count}</strong></li>\n'
            elif risk_level == 'MEDIUM':
                html += f'<li><strong style="color: #ffcc00;">🟡 MEDIUM: {count}</strong></li>\n'
            elif risk_level == 'LOW':
                html += f'<li>🟢 LOW: {count}</li>\n'
            else:
                html += f'<li>⚪ {risk_level}: {count}</li>\n'
        
        html += """
                        </ul>
                    </div>
                    
                    <div class="stats">
                        <h3>Events by Service</h3>
                        <ul>
        """
        
        for service, count in sorted(services.items(), key=lambda x: x[1], reverse=True)[:10]:
            html += f"<li>{service}: {count}</li>\n"
        
        html += """
                        </ul>
                    </div>
                    
                    <div class="stats">
                        <h3>Events by Category</h3>
                        <ul>
        """
        
        for category, count in sorted(categories.items(), key=lambda x: x[1], reverse=True):
            html += f"<li>{category}: {count}</li>\n"
        
        html += """
                        </ul>
                    </div>
                    
                    <div class="stats">
                        <h3>Events by Region</h3>
                        <ul>
        """
        
        for region, count in sorted(regions.items(), key=lambda x: x[1], reverse=True)[:10]:
            html += f"<li>{region}: {count}</li>\n"
        
        html += f"""
                        </ul>
                    </div>
                </div>
                
                <div class="summary">
                    <h2>Download Full Report</h2>
                    <p>Click the button below to download the complete Excel report with all event details:</p>
                    <a href="{presigned_url}" class="download-button">Download Excel Report</a>
                    <p><small>This link will expire in 7 days.</small></p>
                </div>
            </div>
        </body>
        </html>
        """
        
        return html
        
    except Exception as e:
        print(f"Error generating HTML summary: {str(e)}")
        traceback.print_exc()
        raise


def generate_master_summary_html(events, presigned_url, attachment_included):
    """
    Generate HTML summary for master email
    Adjust messaging based on attachment vs link-only
    """
    try:
        # Calculate statistics
        summary_data = generate_master_summary_data(events)
        
        # Current date
        current_date = datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')
        
        # Attachment messaging
        if attachment_included:
            attachment_msg = """
                <p>The Excel report is attached to this email.</p>
                <p>You can also download it here: <a href="{url}">AWS Health Events Report</a></p>
                <p><small>This link will expire in 7 days.</small></p>
            """.format(url=presigned_url)
        else:
            attachment_msg = """
                <p><strong>Due to the report size, the Excel file is only available via the download link below.</strong></p>
                <p>Download the full report: <a href="{url}">AWS Health Events Report</a></p>
                <p><small>This link will expire in 7 days.</small></p>
            """.format(url=presigned_url)
        
        # Build HTML
        html = f"""
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; }}
                .header {{ background-color: #232F3E; color: white; padding: 20px; }}
                .content {{ padding: 20px; }}
                .summary {{ margin: 20px 0; }}
                .stats {{ background-color: #f5f5f5; padding: 15px; border-radius: 5px; margin: 10px 0; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>AWS Health Events Summary Report [MASTER]</h1>
                <p>Generated: {current_date}</p>
            </div>
            <div class="content">
                <div class="summary">
                    <h2>Summary</h2>
                    <p><strong>Total Open/Upcoming Events:</strong> {summary_data['total_events']}</p>
                    <p><strong>Total Accounts:</strong> {summary_data['total_accounts']}</p>
                    
                    <div class="stats">
                        <h3>Events by Risk Level</h3>
                        <ul>
        """
        
        # Sort risk levels by severity
        risk_order = {'CRITICAL': 0, 'HIGH': 1, 'MEDIUM': 2, 'LOW': 3, 'Unknown': 4}
        for risk_level, count in sorted(summary_data['events_by_risk'].items(), 
                                       key=lambda x: risk_order.get(x[0], 99)):
            if risk_level == 'CRITICAL':
                html += f'<li><strong style="color: #ff4444;">🔴 CRITICAL: {count}</strong></li>\n'
            elif risk_level == 'HIGH':
                html += f'<li><strong style="color: #ff9900;">🟠 HIGH: {count}</strong></li>\n'
            elif risk_level == 'MEDIUM':
                html += f'<li><strong style="color: #ffcc00;">🟡 MEDIUM: {count}</strong></li>\n'
            elif risk_level == 'LOW':
                html += f'<li>🟢 LOW: {count}</li>\n'
            else:
                html += f'<li>⚪ {risk_level}: {count}</li>\n'
        
        html += """
                        </ul>
                    </div>
                    
                    <div class="stats">
                        <h3>Events by Service</h3>
                        <ul>
        """
        
        for service, count in sorted(summary_data['events_by_service'].items(), 
                                    key=lambda x: x[1], reverse=True)[:10]:
            html += f"<li>{service}: {count}</li>\n"
        
        html += """
                        </ul>
                    </div>
                    
                    <div class="stats">
                        <h3>Events by Category</h3>
                        <ul>
        """
        
        for category, count in sorted(summary_data['events_by_category'].items(), 
                                     key=lambda x: x[1], reverse=True):
            html += f"<li>{category}: {count}</li>\n"
        
        html += """
                        </ul>
                    </div>
                    
                    <div class="stats">
                        <h3>Events by Region</h3>
                        <ul>
        """
        
        for region, count in sorted(summary_data['events_by_region'].items(), 
                                   key=lambda x: x[1], reverse=True)[:10]:
            html += f"<li>{region}: {count}</li>\n"
        
        html += f"""
                        </ul>
                    </div>
                </div>
                
                <div class="summary">
                    <h2>Download Report</h2>
                    {attachment_msg}
                </div>
            </div>
        </body>
        </html>
        """
        
        return html
        
    except Exception as e:
        print(f"Error generating HTML summary: {str(e)}")
        traceback.print_exc()
        raise


def send_master_email_with_attachment(html_content, presigned_url, excel_bytes):
    """
    Send master email with Excel file attached using SendRawEmail API
    """
    try:
        print(f"Sending master email with attachment to: {MASTER_RECIPIENT_EMAIL}")
        
        # Create email subject
        current_date = datetime.now(timezone.utc).strftime('%Y-%m-%d')
        subject = f"AWS Health Events Summary [MASTER] - {current_date}"
        
        # Create MIME message
        msg = MIMEMultipart('mixed')
        msg['Subject'] = subject
        msg['From'] = SENDER_EMAIL
        msg['To'] = MASTER_RECIPIENT_EMAIL
        
        # Add HTML body
        body = MIMEMultipart('alternative')
        html_part = MIMEText(html_content, 'html')
        body.attach(html_part)
        msg.attach(body)
        
        # Add Excel attachment
        timestamp = datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')
        filename = f"AWS_Health_Events_Master_{timestamp}.xlsx"
        attachment = MIMEApplication(excel_bytes)
        attachment.add_header('Content-Disposition', 'attachment', filename=filename)
        msg.attach(attachment)
        
        # Send via SES
        response = ses_client.send_raw_email(
            Source=SENDER_EMAIL,
            Destinations=[MASTER_RECIPIENT_EMAIL],
            RawMessage={
                'Data': msg.as_string()
            }
        )
        
        print(f"Master email with attachment sent successfully. MessageId: {response['MessageId']}")
        
    except Exception as e:
        print(f"Error sending master email with attachment: {str(e)}")
        traceback.print_exc()
        raise


def send_master_email_link_only(html_content, presigned_url):
    """
    Send master email without attachment (link only) using SendEmail API
    """
    try:
        print(f"Sending master email (link only) to: {MASTER_RECIPIENT_EMAIL}")
        
        # Create email subject
        current_date = datetime.now(timezone.utc).strftime('%Y-%m-%d')
        subject = f"AWS Health Events Summary [MASTER] - {current_date}"
        
        # Send email using SES
        response = ses_client.send_email(
            Source=SENDER_EMAIL,
            Destination={
                'ToAddresses': [MASTER_RECIPIENT_EMAIL]
            },
            Message={
                'Subject': {
                    'Data': subject,
                    'Charset': 'UTF-8'
                },
                'Body': {
                    'Html': {
                        'Data': html_content,
                        'Charset': 'UTF-8'
                    }
                }
            }
        )
        
        print(f"Master email (link only) sent successfully. MessageId: {response['MessageId']}")
        
    except Exception as e:
        print(f"Error sending master email (link only): {str(e)}")
        traceback.print_exc()
        raise


def send_no_events_email():
    """
    Send email when no open events are found
    """
    try:
        current_date = datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')
        
        html_content = f"""
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; }}
                .header {{ background-color: #232F3E; color: white; padding: 20px; }}
                .content {{ padding: 20px; }}
                .success {{ background-color: #d4edda; padding: 15px; border-radius: 5px; color: #155724; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>AWS Health Events Summary Report</h1>
                <p>Generated: {current_date}</p>
            </div>
            <div class="content">
                <div class="success">
                    <h2>✓ No Open Health Events</h2>
                    <p>There are currently no open or upcoming AWS Health events to report.</p>
                </div>
            </div>
        </body>
        </html>
        """
        
        subject = f"AWS Health Events Summary - No Open Events - {datetime.now(timezone.utc).strftime('%Y-%m-%d')}"
        
        ses_client.send_email(
            Source=SENDER_EMAIL,
            Destination={
                'ToAddresses': [MASTER_RECIPIENT_EMAIL]
            },
            Message={
                'Subject': {
                    'Data': subject,
                    'Charset': 'UTF-8'
                },
                'Body': {
                    'Html': {
                        'Data': html_content,
                        'Charset': 'UTF-8'
                    }
                }
            }
        )
        
        print("No events email sent successfully")
        
    except Exception as e:
        print(f"Error sending no events email: {str(e)}")
        traceback.print_exc()
        raise


# ============================================================================
# Per-Account Email Functions
# ============================================================================

def fetch_open_events():
    """
    Query DynamoDB for health events where statusCode != "closed"
    Returns: list of open health events
    """
    try:
        table = dynamodb.Table(DYNAMODB_HEALTH_EVENTS_TABLE_NAME)
        
        print("Scanning DynamoDB for open events (excluding closed)...")
        
        # Scan for all events excluding closed ones
        response = table.scan(
            FilterExpression='attribute_not_exists(statusCode) OR #status <> :closed',
            ExpressionAttributeNames={
                '#status': 'statusCode'
            },
            ExpressionAttributeValues={
                ':closed': 'closed'
            }
        )
        
        events = response.get('Items', [])
        
        # Handle pagination
        while 'LastEvaluatedKey' in response:
            response = table.scan(
                FilterExpression='attribute_not_exists(statusCode) OR #status <> :closed',
                ExpressionAttributeNames={
                    '#status': 'statusCode'
                },
                ExpressionAttributeValues={
                    ':closed': 'closed'
                },
                ExclusiveStartKey=response['LastEvaluatedKey']
            )
            events.extend(response.get('Items', []))
        
        closed_count = 0
        # Count closed events for logging
        response_all = table.scan()
        all_events = response_all.get('Items', [])
        while 'LastEvaluatedKey' in response_all:
            response_all = table.scan(ExclusiveStartKey=response_all['LastEvaluatedKey'])
            all_events.extend(response_all.get('Items', []))
        
        closed_count = sum(1 for e in all_events if e.get('statusCode') == 'closed')
        
        print(f"Fetched {len(events)} open events (excluded {closed_count} closed events)")
        return events
        
    except Exception as e:
        print(f"Error fetching open events: {str(e)}")
        traceback.print_exc()
        raise


def fetch_custom_email_mappings():
    """
    Scan account-email-mappings DynamoDB table
    Returns: dict mapping accountId -> email
    Handle table not found gracefully
    """
    try:
        if not ACCOUNT_EMAIL_MAPPINGS_TABLE:
            print("Account email mappings table not configured, skipping custom mappings")
            return {}
        
        print(f"Scanning account email mappings table: {ACCOUNT_EMAIL_MAPPINGS_TABLE}")
        
        table = dynamodb.Table(ACCOUNT_EMAIL_MAPPINGS_TABLE)
        
        # Scan the entire table
        response = table.scan()
        items = response.get('Items', [])
        
        # Handle pagination
        while 'LastEvaluatedKey' in response:
            response = table.scan(ExclusiveStartKey=response['LastEvaluatedKey'])
            items.extend(response.get('Items', []))
        
        # Build mapping dict
        mappings = {}
        for item in items:
            account_id = item.get('accountId')
            email = item.get('email')
            if account_id and email:
                mappings[account_id] = email
        
        print(f"Fetched {len(mappings)} custom email mappings")
        return mappings
        
    except dynamodb.meta.client.exceptions.ResourceNotFoundException:
        print(f"Account email mappings table not found: {ACCOUNT_EMAIL_MAPPINGS_TABLE}")
        return {}
    except Exception as e:
        print(f"Error fetching custom email mappings (continuing with Organizations only): {str(e)}")
        traceback.print_exc()
        return {}


def fetch_account_owners():
    """
    Call AWS Organizations ListAccounts API
    Call DescribeAccount for each account to get email
    Returns: dict mapping accountId -> {accountName, ownerEmail}
    """
    try:
        print("Fetching account owners from AWS Organizations...")
        
        # List all accounts in the organization
        accounts = []
        paginator = organizations_client.get_paginator('list_accounts')
        
        for page in paginator.paginate():
            accounts.extend(page.get('Accounts', []))
        
        print(f"Found {len(accounts)} accounts in organization")
        
        # Build mapping dict with account details
        account_mappings = {}
        for account in accounts:
            account_id = account.get('Id')
            account_name = account.get('Name', 'Unknown')
            owner_email = account.get('Email', '')
            status = account.get('Status', '')
            
            # Only include active accounts
            if status == 'ACTIVE' and account_id and owner_email:
                account_mappings[account_id] = {
                    'accountName': account_name,
                    'ownerEmail': owner_email
                }
        
        print(f"Fetched {len(account_mappings)} active account owners")
        return account_mappings
        
    except Exception as e:
        print(f"Error fetching account owners from Organizations: {str(e)}")
        traceback.print_exc()
        # Return empty dict to allow processing to continue
        return {}


def merge_email_mappings(custom_mappings, org_mappings):
    """
    Merge custom mappings with Organizations data
    Prioritize custom mappings over Organizations
    Returns: unified mapping dict
    """
    try:
        print("Merging custom mappings with Organizations data...")
        
        # Start with Organizations mappings
        merged = {}
        for account_id, account_info in org_mappings.items():
            merged[account_id] = {
                'accountName': account_info['accountName'],
                'ownerEmail': account_info['ownerEmail'],
                'mappingSource': 'AWS Organizations'
            }
        
        # Override with custom mappings where they exist
        for account_id, custom_email in custom_mappings.items():
            if account_id in merged:
                # Update existing account with custom email
                merged[account_id]['ownerEmail'] = custom_email
                merged[account_id]['mappingSource'] = 'DynamoDB mapping'
            else:
                # Account not in Organizations (edge case), add it anyway
                merged[account_id] = {
                    'accountName': f'Account-{account_id}',
                    'ownerEmail': custom_email,
                    'mappingSource': 'DynamoDB mapping'
                }
        
        custom_count = sum(1 for v in merged.values() if v['mappingSource'] == 'DynamoDB mapping')
        org_count = len(merged) - custom_count
        
        print(f"Merged mappings: {len(merged)} total ({custom_count} custom, {org_count} from Organizations)")
        return merged
        
    except Exception as e:
        print(f"Error merging email mappings: {str(e)}")
        traceback.print_exc()
        raise


def build_email_mappings_info(all_accounts, custom_mappings, org_mappings):
    """
    Build metadata for Account Email Mapping spreadsheet
    Determine mappingSource and availabilityStatus for each account
    Returns: list of dicts with account mapping info
    """
    try:
        print("Building email mappings info for spreadsheet...")
        
        mappings_info = []
        
        for account_id, account_info in all_accounts.items():
            account_name = account_info['accountName']
            email_address = account_info['ownerEmail']
            mapping_source = account_info['mappingSource']
            
            # Determine availability status based on presence in both sources
            in_custom = account_id in custom_mappings
            in_org = account_id in org_mappings
            
            if in_custom and in_org:
                availability_status = 'Organizations + DynamoDB'
            elif in_custom:
                availability_status = 'DynamoDB only'
            elif in_org:
                availability_status = 'Organizations only'
            else:
                availability_status = 'Unknown'
            
            mappings_info.append({
                'accountId': account_id,
                'accountName': account_name,
                'emailAddress': email_address,
                'mappingSource': mapping_source,
                'availabilityStatus': availability_status
            })
        
        # Sort by account ID for consistent ordering
        mappings_info.sort(key=lambda x: x['accountId'])
        
        print(f"Built email mappings info for {len(mappings_info)} accounts")
        return mappings_info
        
    except Exception as e:
        print(f"Error building email mappings info: {str(e)}")
        traceback.print_exc()
        raise


def group_events_by_account(events):
    """
    Group health events by accountId
    Returns: dict mapping accountId -> list of events
    """
    try:
        print("Grouping events by account...")
        
        grouped = defaultdict(list)
        
        for event in events:
            account_id = event.get('accountId')
            if account_id:
                grouped[account_id].append(event)
        
        print(f"Grouped {len(events)} events into {len(grouped)} accounts")
        return dict(grouped)
        
    except Exception as e:
        print(f"Error grouping events by account: {str(e)}")
        traceback.print_exc()
        raise


def consolidate_accounts_by_email(account_events, email_mappings):
    """
    Group accounts by destination email address
    Aggregate events from multiple accounts mapped to same email
    Include emailMappingsInfo for each account
    Returns: dict mapping email -> {accountIds, accountNames, eventKeys, mappingsInfo}
    Note: eventKeys contains only {eventArn, accountId} to keep SQS message size small
    """
    try:
        print("Consolidating accounts by email address...")
        
        consolidated = defaultdict(lambda: {
            'accountIds': [],
            'accountNames': [],
            'eventKeys': [],  # Changed from 'events' to 'eventKeys'
            'mappingsInfo': []
        })
        
        for account_id, events in account_events.items():
            # Get email mapping for this account
            if account_id in email_mappings:
                email = email_mappings[account_id]['ownerEmail']
                account_name = email_mappings[account_id]['accountName']
                mapping_source = email_mappings[account_id]['mappingSource']
                
                # Add to consolidated data
                consolidated[email]['accountIds'].append(account_id)
                consolidated[email]['accountNames'].append(account_name)
                
                # Extract only event keys (eventArn, accountId) to minimize SQS message size
                for event in events:
                    event_key = {
                        'eventArn': event.get('eventArn'),
                        'accountId': event.get('accountId')
                    }
                    consolidated[email]['eventKeys'].append(event_key)
                
                consolidated[email]['mappingsInfo'].append({
                    'accountId': account_id,
                    'accountName': account_name,
                    'emailAddress': email,
                    'mappingSource': mapping_source
                })
            else:
                print(f"Warning: Account {account_id} not found in email mappings, skipping")
        
        # Convert to regular dict
        result = dict(consolidated)
        
        # Count consolidated vs single-account emails
        consolidated_count = sum(1 for v in result.values() if len(v['accountIds']) > 1)
        single_count = len(result) - consolidated_count
        
        print(f"Consolidated into {len(result)} unique emails ({consolidated_count} consolidated, {single_count} single-account)")
        return result
        
    except Exception as e:
        print(f"Error consolidating accounts by email: {str(e)}")
        traceback.print_exc()
        raise


def send_account_email_messages(consolidated_data):
    """
    Send SQS messages for each unique email with consolidated account data
    For each unique email with events, create SQS message
    Include accountIds, accountNames, ownerEmail, isConsolidated, eventKeys, emailMappingsInfo
    Note: Sends only event keys (eventArn, accountId) to keep message size under 256KB limit
    """
    try:
        if not ACCOUNT_EMAIL_QUEUE_URL:
            print("Account email queue URL not configured, skipping account email messages")
            return
        
        print(f"Sending {len(consolidated_data)} account email messages to SQS...")
        
        success_count = 0
        failure_count = 0
        
        for email, data in consolidated_data.items():
            try:
                # Determine if this is a consolidated email
                is_consolidated = len(data['accountIds']) > 1
                
                # Create SQS message with event keys only (not full events)
                message = {
                    'accountIds': data['accountIds'],
                    'accountNames': data['accountNames'],
                    'ownerEmail': email,
                    'isConsolidated': is_consolidated,
                    'eventKeys': data['eventKeys'],  # Changed from 'events' to 'eventKeys'
                    'emailMappingsInfo': data['mappingsInfo'],
                    'timestamp': datetime.now(timezone.utc).isoformat()
                }
                
                # Convert Decimal types to standard Python types for JSON serialization
                message = convert_decimal_to_number(message)
                
                # Log message size for monitoring
                message_size = len(json.dumps(message))
                print(f"SQS message size for {email}: {message_size} bytes ({len(data['eventKeys'])} event keys)")
                
                # Send to SQS
                response = sqs_client.send_message(
                    QueueUrl=ACCOUNT_EMAIL_QUEUE_URL,
                    MessageBody=json.dumps(message)
                )
                
                account_list = ', '.join(data['accountIds'])
                print(f"Sent SQS message for {email} (accounts: {account_list}), MessageId: {response['MessageId']}")
                success_count += 1
                
            except Exception as e:
                print(f"Error sending SQS message for {email}: {str(e)}")
                traceback.print_exc()
                failure_count += 1
        
        print(f"Account email messages sent: {success_count} success, {failure_count} failed")
        
    except Exception as e:
        print(f"Error in send_account_email_messages: {str(e)}")
        traceback.print_exc()
        raise


def process_per_account_emails():
    """
    Process per-account email notifications
    Check ENABLE_PER_ACCOUNT_EMAILS configuration
    Skip per-account email processing if false
    """
    try:
        # Check if per-account emails are enabled
        if not ENABLE_PER_ACCOUNT_EMAILS:
            print("Per-account emails disabled (ENABLE_PER_ACCOUNT_EMAILS=false), skipping")
            return
        
        print("Processing per-account emails...")
        
        # Fetch open events (excluding closed)
        events = fetch_open_events()
        
        if not events:
            print("No open events found for per-account emails")
            return
        
        # Fetch custom email mappings
        custom_mappings = fetch_custom_email_mappings()
        
        # Fetch account owners from Organizations
        org_mappings = fetch_account_owners()
        
        if not org_mappings:
            print("No account owners found from Organizations, skipping per-account emails")
            return
        
        # Merge custom mappings with Organizations data
        all_accounts = merge_email_mappings(custom_mappings, org_mappings)
        
        # Build email mappings info for spreadsheet
        email_mappings_info = build_email_mappings_info(all_accounts, custom_mappings, org_mappings)
        
        # Group events by account
        account_events = group_events_by_account(events)
        
        # Consolidate accounts by email
        consolidated_data = consolidate_accounts_by_email(account_events, all_accounts)
        
        if not consolidated_data:
            print("No accounts with events to send emails for")
            return
        
        # Send SQS messages for each unique email
        send_account_email_messages(consolidated_data)
        
        print(f"Per-account email processing complete: {len(consolidated_data)} emails queued")
        
    except Exception as e:
        print(f"Error processing per-account emails (continuing with master email): {str(e)}")
        traceback.print_exc()
        # Don't raise - we want master email to succeed even if per-account fails
